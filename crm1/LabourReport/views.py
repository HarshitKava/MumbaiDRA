from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages,auth
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from .models import *
from ProductivityReport.models import *
from .forms import *#SiteEngDayForm,UpdateForm,CreateUserForm,Area_Input,AddCont,Add_Labour,Add_Lab_To_Contractor,ResetPasswordForm
from ProductivityReport.forms import *
from datetime import date
from .decorators import allowed_users, unauthenticated_user
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font,Alignment,PatternFill,Border,Side
from datetime import datetime, timedelta
import pandas as pd
from django.core.mail import EmailMessage
from io import BytesIO

def Navbar(request):
    return render(request,'LabourReport/Navbar.html')

@unauthenticated_user
def LoginPage(request):
    if request.method == 'POST':
        userid = request.POST['userid']
        password = request.POST['pass']
        user_auth = authenticate(username=userid, password=password)
        admin = Group.objects.get(name="Admin").user_set.all()
        SE = Group.objects.get(name="Site Engineer").user_set.all()
        SLI = Group.objects.get(name="Site Labour Incharge").user_set.all()
        CLI = Group.objects.get(name="Camp Labour Incharge").user_set.all()
        Mang = Group.objects.get(name="Management").user_set.all()
        if user_auth is not None:
            if user_auth in SE:
                login(request,user_auth)
                return redirect('HomeSE')
            elif user_auth in admin:
                login(request,user_auth)
                return redirect('HomeAdmin')
            elif user_auth in SLI:
                login(request,user_auth)
                print("in")
                return redirect('HomeSLI')
            elif user_auth in CLI:
                login(request,user_auth)
                return redirect('HomeCLI')
            elif user_auth in Mang:
                login(request,user_auth)
                return redirect('HomeMang')
        else:
            messages.info(request,'Credential is incorrect')
            return redirect('/')
    return render(request,'LabourReport/LoginPage.html')

def LogoutUser(request):
    logout(request)
    return redirect('Login')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def HomeSE(request):
    return render(request,'LabourReport/HomeSE.html')

def CheckStatusAllArea(shift):
    # Get all AreaName
    AreaName = Area.objects.all().order_by('AreaName')
    print("AreaName",AreaName)
    # Convert to Dataframe
    area_df=pd.DataFrame(AreaName.values())
    area_df = area_df[['AreaName']]
    # get unique AreaName
    Area_lst = area_df['AreaName'].unique()
    Area_lst = list(Area_lst)
    # Remove Bhopal and Bangalore
    if 'Bhopal' in Area_lst:
        Area_lst.remove('Bhopal')
    if 'Bangalore' in Area_lst:
        Area_lst.remove('Bangalore')

    # Get all Report_Status
    # Check if Report_Status is present for all AreaName
    # If not present then create Report_Status for that AreaName
    for i in Area_lst:
        Repo_Status = Report_Status.objects.filter(Area=i)
        if Repo_Status.count()==0:
            dic = {
                'Area':i,
                'Date':datetime.now().strftime("%Y-%m-%d"),
                'Status_Day':False,
                'Status_Night':False,
            }
            Report_Status.objects.create(**dic)

    TrueStatus = Report_Status.objects.filter(Date=datetime.now().strftime("%Y-%m-%d"),Status_Day=True).values_list('Area',flat=True)
    if len(TrueStatus)==len(Area_lst):
        return True
    else:
        return False

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def AddDaySE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    # print("d",d1,d2)
    Report=SiteEngDay.objects.filter(created_at__range=[d1,d2],Areaname=Areaname_id).order_by('ContractorName')
    form=SiteEngDayForm()
    if request.method =='POST':
        form=SiteEngDayForm(request.POST)
        print("Request : ",request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddDaySE')
    Repo_Status = Report_Status.objects.filter(Area=Areaname,Date=d1)
    if Repo_Status.count()==0:
        Report_Status.objects.update_or_create(Area=Areaname,Date=d1,defaults={'Status_Day':False,'Status_Night':False})
    Repo_Status = Report_Status.objects.filter(Area=Areaname,Date=d1)


    # if CheckStatusAllArea():
    #     print("Status_Day",Repo_Status[0].Status_Day)
    #     Repo_Status = Report_Status.objects.filter(Area=Areaname)
    #     Repo_Status.update(Status_Day=True)
    #     print("Status_Day",Repo_Status[0].Status_Day)

    return render(request,'LabourReport/SiteEngAddDayData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id,'Status':Repo_Status[0].Status_Day})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def EditDaySE(request,i):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    data=SiteEngDay.objects.get(id=i)
    dic = {
        'Areaname':data.Areaname.id,
        'ContractorName':data.ContractorName.id,
        'LabourCategory':data.LabourCategory.id,
        'CategoryName':data.CategoryName.id,
        'StructureName':data.StructureName.id,
        'NoLabor':data.NoLabor,
    }
    form=SiteEngDayForm()
    if request.method =='POST':
        print('request',request.POST)
        form=SiteEngDayForm(request.POST)
        # update on id = data.id

        if form.is_valid():
            my_data = SiteEngDay.objects.get(id=data.id)
            my_data.Areaname = Area.objects.get(id=request.POST['Areaname'])
            my_data.ContractorName = ContractorDetail.objects.get(id=request.POST['ContractorName'])
            my_data.LabourCategory = LabourOfContractor.objects.get(id=request.POST['LabourCategory'])
            my_data.CategoryName = CategoryOfDeployment.objects.get(id=request.POST['CategoryName'])
            my_data.StructureName = Structure.objects.get(id=request.POST['StructureName'])
            my_data.NoLabor = request.POST['NoLabor']
            my_data.save()
        return redirect('AddDaySE')
    return render(request,'LabourReport/EditSiteEngDay.html',{'form':form,'dict':dic,'Areaname_id':Areaname_id,'labor':data.NoLabor})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def ViewDaySE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2,type(d1))
    Report=SiteEngDay.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/SiteEngViewDayData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def ViewNightSE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SiteEngNight.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/SiteEngViewNightData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def AddNightSE(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SiteEngNight.objects.filter(created_at__range=[d1,d2],Areaname=Areaname_id)
    form=SiteEngNightForm()
    if request.method =='POST':
        form=SiteEngNightForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('/AddNightSE/')
    return render(request,'LabourReport/SiteEngAddNightData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def DeleteDaySE(request,i):
    new=SiteEngDay.objects.get(id=i)
    new.delete()
    return redirect('/AddDaySE/')


@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def DeleteNightSE(request,i):
    new=SiteEngNight.objects.get(id=i)
    new.delete()
    return redirect('/AddNightSE/')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def DLRSummary(request):
    if request.method=='POST':
        # print(request.POST)
        query_dict = request.POST.dict()
        print(query_dict.keys())
        if 'Show' in query_dict.keys():
            # Get data from SiteEngDay table
            current_user = request.user
            Areaname = Area.objects.filter(Username=current_user.username)
            d1 = query_dict['From']
            d2 = query_dict['To']
            print(query_dict['Shift'])
            if query_dict['Shift']=='Day':
                data = SiteEngDay.objects.filter(created_at__range=[d1,d2],Areaname=Areaname[0].id).order_by('ContractorName')
            else:
                data = SiteEngNight.objects.filter(created_at__range=[d1,d2],Areaname=Areaname[0].id).order_by('ContractorName')
            df = pd.DataFrame(columns=['ContractorName','Date','LabourCategory','CategoryName','NoLabor'])
            for i in data:
                # print all data
                date = str(i.created_at)[:10]
                date = date[8:10]+"/"+date[5:7]+"/"+date[0:4]
                df.loc[len(df)] = {'ContractorName':i.ContractorName,'Date':date,'LabourCategory':i.LabourCategory,'CategoryName':i.CategoryName,'NoLabor':i.NoLabor}
                # df = df.append({'ContractorName':i.ContractorName,'Date':date,'LabourCategory':i.LabourCategory,'CategoryName':i.CategoryName,'NoLabor':i.NoLabor},ignore_index=True)
            # Get unique ContractorName from df
            cont_name = df['ContractorName'].unique()
            cont_row_span = {}
            for i in cont_name:
                cont_row_span[i] = len(df[df['ContractorName']==i])
            return render(request,'LabourReport/DLR_Summary.html',{'df':df,'from':d1,'to':d2,'Shift':query_dict['Shift'],'cont_row_span':cont_row_span})
        elif 'Export' in query_dict.keys():
            current_user = request.user
            Areaname = Area.objects.filter(Username=current_user.username)
            d1 = query_dict['From']
            d2 = query_dict['To']
            print(query_dict['Shift'])
            if query_dict['Shift']=='Day':
                data = SiteEngDay.objects.filter(created_at__range=[d1,d2],Areaname=Areaname[0].id).order_by('ContractorName')
            else:
                data = SiteEngNight.objects.filter(created_at__range=[d1,d2],Areaname=Areaname[0].id).order_by('ContractorName')
            df = pd.DataFrame(columns=['ContractorName','Date','LabourCategory','CategoryName','NoLabor'])
            for i in data:
                # print all data
                date = str(i.created_at)[:10]
                date = date[8:10]+"/"+date[5:7]+"/"+date[0:4]
                print(date)
                df.loc[len(df)] = {'ContractorName':i.ContractorName,'Date':date,'LabourCategory':i.LabourCategory,'CategoryName':i.CategoryName,'NoLabor':i.NoLabor}
                # df = df.append({'ContractorName':i.ContractorName,'Date':date,'LabourCategory':i.LabourCategory,'CategoryName':i.CategoryName,'NoLabor':i.NoLabor},ignore_index=True)
            # Create Workbook and add worksheet
            workbook = Workbook()
            # Get active worksheet/tab
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename="DLR "+d1+" to "+d2+" "+query_dict['Shift']+".xlsx"
            response['Content-Disposition'] = 'attachment; filename='+filename
            worksheet = workbook.active
            worksheet.title = 'DLR Summary'

            # Define the titles for columns
            columns = [
                'ContractorName',
                'Date',
                'LabourCategory',
                'CategoryName',
                'NoLabor',
            ]

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            # Bold the header
            for cell in worksheet["1:1"]:
                cell.font = Font(bold=True)
            row_num = 2
            # Append data in sheet
            for i in range(0,len(df)):
                # print(df.iloc[i])
                print(df.iloc[i]['ContractorName'])
                # Reverse the date format
                date = str(df.iloc[i]['Date'])
                row = [
                    str(df.iloc[i]['ContractorName']),
                    date,
                    str(df.iloc[i]['LabourCategory']),
                    str(df.iloc[i]['CategoryName']),
                    df.iloc[i]['NoLabor'],
                ]
                print(row)
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                row_num += 1
            # Save the file
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length+3
                worksheet.column_dimensions[column].width = adjusted_width
            workbook.save(response)
            return response

    return render(request,'LabourReport/DLR_Summary.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def HomeAdmin(request):
    # Get Data of users
    User_data=User.objects.all().order_by('username')
    User_data=User.objects.all().order_by('username')
    df = pd.DataFrame(columns=['username','email','group','id'])
    for i in User_data:
        if i.username != 'harshitkava' and i.username != 'HarhsitKava':
            # print(i)
            df.loc[len(df)] = {'username':i.username,'id':i.id,'email':i.email,'group':i.groups.all()[0]}

    Area_data=Area.objects.all().order_by('AreaName')
    # Convert to Dataframe
    df1=pd.DataFrame(Area_data.values())
    df1 = pd.DataFrame(columns=['AreaName','Username'])
    # df1 = df1[['AreaName','Username']]
    # Merge Dataframes
    df2 = pd.merge(df,df1,how='left',left_on='username',right_on='Username')
    # drop Username column
    df2 = df2.drop(['Username'],axis=1)
    return render(request,'LabourReport/Admin/HomeAd.html',{'data':df2})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def AddWorkArea(request):
    Form=WorkAreaForm()
    data = WorkArea.objects.all()
    if request.method=='POST':
        Form=WorkAreaForm(request.POST)
        if Form.is_valid():
            Form.save()
            return redirect('AddWorkArea')
    return render(request,'LabourReport/Admin\WorkArea.html',{'Form':Form,'data':data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditWorkArea(request,i):
    i = float(i)
    data = WorkArea.objects.get(pk=i)
    form = WorkAreaForm(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('ShowWorkArea')
    return render(request,'LabourReport/Admin/EditWorkArea.html',{'Form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteWorkArea(request,i):
    i = float(i)
    data = WorkArea.objects.get(pk=i)
    data.delete()
    return redirect('ShowWorkArea')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowWorkArea(request):
    data = WorkArea.objects.all().order_by('WorkAreaName')
    df = pd.DataFrame(columns=['WorkAreaName','id'])
    for i in data:
        df.loc[len(df)] = {'WorkAreaName':i.WorkAreaName,'id':i.id}
        # df = df.append({'WorkAreaName':i.WorkAreaName,'id':i.id},ignore_index=True)
    return render(request,'LabourReport/Admin/ShowWorkArea.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditUser(request,i):
    # Get Data of users
    i = float(i)
    User_data=User.objects.get(pk=i)
    Area_data=Area.objects.get(Username=User_data.username)
    registration_form=CreateUserForm(request.POST or None,instance=User_data)
    Area_input=Area_Input(request.POST or None,instance=Area_data)
    print("*************",registration_form.is_valid())
    if registration_form.is_valid() and Area_input.is_valid():
        registration_form.save()
        Area_input.save()
        print(1)
        return redirect('HomeAdmin')
    return render(request,'LabourReport/Admin/EditUser.html',{'form1':registration_form,'form2':Area_input})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteUser(request,i):
    i = float(i)
    User_data=User.objects.get(pk=i)
    User_data.delete()
    return redirect('HomeAdmin')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowContractor(request):
    data = ContractorDetail.objects.all().order_by('ContractorName')
    df = pd.DataFrame(columns=['ContractorName','ContractorNumber','id'])
    for i in data:
        df.loc[len(df)] = {'ContractorName':i.ContractorName,'ContractorNumber':i.ContractorNumber,'id':i.id}
        # df = df.append({'ContractorName':i.ContractorName,'ContractorNumber':i.ContractorNumber,'id':i.id},ignore_index=True)
    return render(request,'LabourReport/Admin/ShowContractor.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditContractor(request,i):
    i = float(i)
    data = ContractorDetail.objects.get(pk=i)
    form = AddCont(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('Contractor')
    return render(request,'LabourReport/Admin/EditContractor.html',{'form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteContractor(request,i):
    i = float(i)
    data = ContractorDetail.objects.get(pk=i)
    data.delete()
    return redirect('Contractor')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowStructure(request):
    data = Structure.objects.all().order_by('StructureName')
    df = pd.DataFrame(columns=['StructureName','id'])
    for i in data:
        df.loc[len(df)] = {'StructureName':i.StructureName,'id':i.id}
        # df = df.append({'StructureName':i.StructureName,'id':i.id},ignore_index=True)
    return render(request,'LabourReport/Admin/ShowStructure.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditStructure(request,i):
    i = float(i)
    data = Structure.objects.get(pk=i)
    form = StructureForm(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('ShowStructure')
    return render(request,'LabourReport/Admin/EditStructure.html',{'Form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteStructure(request,i):
    i = float(i)
    data = Structure.objects.get(pk=i)
    data.delete()
    return redirect('ShowStructure')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowLabours(request):
    data = AddLabour.objects.all().order_by('LabourCategory')
    df = pd.DataFrame(columns=['LabourCategory','id'])
    for i in data:
        df.loc[len(df)] = {'LabourCategory':i.LabourCategory,'id':i.id}
        # df = df.append({'LabourName':i.LabourName,'id':i.id},ignore_index=True)
    return render(request,'LabourReport/Admin/ShowLabours.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditLabours(request,i):
    i = float(i)
    data = AddLabour.objects.get(pk=i)
    form = Add_Labour(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('ShowLabours')
    return render(request,'LabourReport/Admin/EditLabours.html',{'form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteLabours(request,i):
    i = float(i)
    data = AddLabour.objects.get(pk=i)
    data.delete()
    return redirect('ShowLabours')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowLabourOfContractor(request):
    data = LabourOfContractor.objects.all()
    df = pd.DataFrame(columns=['ContractorName','LabourCategory','id'])
    for i in data:
        df.loc[len(df)] = {'ContractorName':str(i.ContractorName),'LabourCategory':i.LabourCategory,'id':i.id}
        # df = df.append({'ContractorName':i.ContractorName,'id':i.id},ignore_index=True)
    df = df.sort_values(by='ContractorName')
    # print(df.sort_values(by='ContractorName',ascending=True))
    return render(request,'LabourReport/Admin/ShowLabourOfContractor.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditLabourOfContractor(request,i):
    i = float(i)
    data = LabourOfContractor.objects.get(pk=i)
    form = Add_Lab_To_Contractor(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('ShowLabourOfContractor')
    return render(request,'LabourReport/Admin/EditLabourOfContractor.html',{'form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteLabourOfContractor(request,i):
    i = float(i)
    data = LabourOfContractor.objects.get(pk=i)
    data.delete()
    return redirect('ShowLabourOfContractor')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ShowActivity(request):
    data = CategoryOfDeployment.objects.all().order_by('ActivityName')
    df = pd.DataFrame(columns=['CategoryName','ActivityName','id'])
    for i in data:
        df.loc[len(df)] = {'CategoryName':i.CategoryName,'ActivityName':i.ActivityName,'id':i.id}
        # df = df.append({'ActivityName':i.ActivityName,'id':i.id},ignore_index=True)
    return render(request,'LabourReport/Admin/ShowActivity.html',{'data':df})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def EditActivity(request,i):
    i = float(i)
    data = CategoryOfDeployment.objects.get(pk=i)
    form = CategoryOfDeploymentForm(request.POST or None,instance=data)
    if form.is_valid():
        form.save()
        return redirect('ShowActivity')
    return render(request,'LabourReport/Admin/EditActivity.html',{'Form':form})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def DeleteActivity(request,i):
    i = float(i)
    data = CategoryOfDeployment.objects.get(pk=i)
    data.delete()
    return redirect('ShowActivity')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def ManagerDashboard(request):
    return render(request,'LabourReport/Management/Dashboard.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def AddUser(request):
    registration_form=CreateUserForm()
    Area_input=Area_Input()
    # descending order of groups in User table
    User_data=User.objects.all().order_by('username')
    if request.method == 'POST':
        registration_form=CreateUserForm(request.POST)
        Area_input=Area_Input(request.POST)

        print(2)
        print(request.POST.get('groups'))
        print(registration_form.errors)
        if registration_form.is_valid():
            print(3)
            user=registration_form.save()
            if request.POST.get('groups')== "1":
                grp = Group.objects.get(name='Site Engineer')
                # grp=list(grp)
                user.groups.add(grp)
            elif request.POST.get('groups')== "2":
                grp = Group.objects.get(name='Site Labour Incharge')
                user.groups.add(grp)
            elif request.POST.get('groups')== "3":
                grp = Group.objects.get(name='Admin')
                user.groups.add(grp)
            elif request.POST.get('groups')== "4":
                grp = Group.objects.get(name='Management')
                user.groups.add(grp)
            elif request.POST.get('groups')== "5":
                grp = Group.objects.get(name='Camp Labour Incharge')
                user.groups.add(grp)
            if Area_input.is_valid():
                Area_input.save()
                print(4)
                return redirect('AddUser')
    return render(request,'LabourReport/Admin/User.html',{'form1':registration_form,'form2':Area_input,'User_data':User_data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def AddContractor(request):
    form=AddCont()
    data = ContractorDetail.objects.all()
    if request.method == 'POST':
        form=AddCont(request.POST)
        if form.is_valid():
            form.save()
        return render(request,'LabourReport/Admin/Contractor.html',{'form':form,'data':data})
    return render(request,'LabourReport/Admin/Contractor.html',{'form':form,'data':data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def AddLabours(request):
    form=Add_Labour()
    labour_data = AddLabour.objects.all()
    if request.method == 'POST':
        form=Add_Labour(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ShowLabours')

    return render(request,'LabourReport/Admin/Labour.html',{'form':form,'labour_data':labour_data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def LaboursOfContractor(request):
    form=Add_Lab_To_Contractor()
    data = LabourOfContractor.objects.all()
    if request.method == 'POST':
        form=Add_Lab_To_Contractor(request.POST)

        if form.is_valid():
            form.save()
            return render(request,'LabourReport/Admin/LabourToCont.html',{'form':form,'data':data})
    return render(request,'LabourReport/Admin/LabourToCont.html',{'form':form,'data':data})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Admin'])
def ResetPassword (request):
    form=ResetPasswordForm()
    if request.method == 'POST':
        form=ResetPasswordForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('HomeAdmin')
    return render(request,'LabourReport/Admin/ResetPassword.html',{'form':form})

@login_required(login_url='Login')
# @allowed_users(allowed_roles=['Site Engineer'])
def load_labour(request):
    contractor_id = request.GET.get('contractor_id')
    if len(contractor_id.split(' ')) > 1:
        val = contractor_id.split(' ')[1]
        val = int(val)
        contractor_id = contractor_id.split(' ')[0]
    else:
        val = 0


    # if request.Get.get('val') is not None:
    # val =request.GET.get('val')
    # else:
    #     val = 0
    contractor_name = ContractorDetail.objects.get(id=contractor_id)
    labourofCont = LabourOfContractor.objects.filter(ContractorName=contractor_name).order_by('LabourCategory')
    labour =AddLabour.objects.filter().order_by('LabourCategory')
    # print(val,type(val))
    return render(request, 'LabourReport/Admin/labour_dropdown_list_options.html', {'val':val,'labour': labour,'labourofCont':labourofCont})

@login_required(login_url='Login')
# @allowed_users(allowed_roles=['Site Engineer'])
def load_cat(request):
    Labour_id = request.GET.get('contractor_id')
    if len(Labour_id.split(' '))>1:
        val = Labour_id.split(' ')[1]
        val = int(val)
        Labour_id = Labour_id.split(' ')[0]
    else:
        val = 0
    print('hi')
    # print(Labour_id)
    # CategoryOfDeployment_id = CategoryOfDeployment.objects.get(id=Labour_id)
    # print(CategoryOfDeployment_id)
    labour_name = LabourOfContractor.objects.get(id=Labour_id)
    # print(labour_name,type(labour_name),str(labour_name))
    labour =AddLabour.objects.get(LabourCategory=labour_name)
    # print(labour,type(labour))
    Category=CategoryOfDeployment.objects.filter(ActivityName=labour)
    # print(Category,type(Category))
    # {'labour': labour,'labourofCont':labourofCont}
    return render(request, 'LabourReport/Admin/category_dropdown_list_options.html',{'val':val,'Category':Category})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def HomeSLI(request):
    return render(request,'LabourReport/SLI/HomeSLI.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def AddDaySLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")

    print("d",d1,d2)
    # get report from 01:00:00 of d1  to 01:00:00 to d2

    Report=SLIDay.objects.filter(created_at__range=[d1,d2])
    # get list of all NoLabor
    NoLabor = SLIDay.objects.filter(created_at__range=[d1,d2]).values_list('NoLabor',flat=True)
    NoLabor = list(NoLabor)
    Nolabor = sum(NoLabor)
    print(Report)
    form=SLIDayForm()
    if request.method =='POST':
        form=SLIDayForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddDaySLI')
    return render(request,'LabourReport/SLI/SLIAddDayData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id,'sum':Nolabor})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def DeleteDaySLI(request,i):
    new=SLIDay.objects.get(id=i)
    new.delete()
    return redirect('AddDaySLI')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def AddNightSLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    d2 = tomorrow.replace(hour=1, minute=0, second=0, microsecond=0)
    d1 = today.replace(hour=1, minute=0, second=0, microsecond=0)

    # print("d1:", d1)
    # print("d2:", d2)
    print("d",d1,d2)
    Report=SLINight.objects.filter(created_at__range=[d1,d2])
    form=SLINightForm()
    if request.method =='POST':
        form=SLINightForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddNightSLI')
    return render(request,'LabourReport/SLI/SLIAddNightData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def DeleteNightSLI(request,i):
    new=SLINight.objects.get(id=i)
    new.delete()
    return redirect('AddNightSLI')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def ViewDaySLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLIDay.objects.filter()
    return render(request,'LabourReport/SLI/SLIViewDayData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Labour Incharge'])
def ViewNightSLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=SLINight.objects.filter()
    return render(request,'LabourReport/SLI/SLIViewNightData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def ViewDayCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLIDay.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/CLI/CLIViewDayData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def ViewNightCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLINight.objects.filter(Areaname=Areaname_id)
    return render(request,'LabourReport/CLI/CLIViewNightData.html',{'Report':Report,'Areaname':Areaname,'Areaname_id':Areaname_id})
    print("Areaname",Areaname)

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def HomeCLI(request):
    return render(request,'LabourReport/CLI/HomeCLI.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def AddDayCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLIDay.objects.filter(created_at__range=[d1,d2])
    form=CLIDayForm()
    if request.method =='POST':
        form=CLIDayForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddDayCLI')
    return render(request,'LabourReport/CLI/CLIAddDayData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def DeleteDayCLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=CLIDay.objects.all()
    new=CLIDay.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=CLIDayForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddDayCLI')
    return render(request,'LabourReport/CLI/CLIDelDayData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def AddNightCLI(request):
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    today = datetime.now()
    tomorrow = today + timedelta(1)
    d2=tomorrow.strftime("%Y-%m-%d")
    d1 = today.strftime("%Y-%m-%d")
    print("d",d1,d2)
    Report=CLINight.objects.filter(created_at__range=[d1,d2])
    form=CLINightForm()
    if request.method =='POST':
        form=CLINightForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('AddNightCLI')
    return render(request,'LabourReport/CLI/CLIAddNightData.html',{'Report':Report,'form':form,'Areaname':Areaname,'Areaname_id':Areaname_id})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Camp Labour Incharge'])
def DeleteNightCLI(request,i,A,N,C,L,H):
    today = date.today()
    Report=CLINight.objects.all()
    new=CLINight.objects.get(id=i)
    updateData={
        'ContName': N,
        'AreaName': A,
        'LaborCat': C,
        'NoLabor': L,
        'NoHelp': H,
    }
    form=CLINightForm(initial=updateData)
    if request.method =='POST':
        new.delete()
        return redirect('AddNightCLI')
    return render(request,'LabourReport/CLI/CLIDelNightData.html',{'Report':Report,'form':form,'today':today,'new':new})

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def LabourRequest(request):
    if  request.method == 'POST':
        print("request",request.POST)
        # return redirect('LabourRequest')
        import time
        # import webbrowser as web
        import pyautogui as pg
        # import pywhatkit
        # import urllib.parse
        ph = ['+919825040957','+919099853457']
        # # ph = urllib.parse.quote(ph)
        # ph1 = ",".join([urllib.parse.quote(p) for p in ph])


        # print(ph1)
        # msg = "Testing Message"
        message = "Hello from Python!"

        for i in range(0,len(ph)):
            # Get the current time
            print("ph :",ph[i])
            # pg.hotkey('ctrl', 't')
            # # write the url and press enter
            # pg.write('https://web.whatsapp.com/send?phone='+ph[i]+'&text='+message)
            # pg.press('enter')
            # time.sleep(10)
            # pg.press('enter')
            # time.sleep(5)
            # pg.hotkey('ctrl', 'w')
            # now = datetime.now()
            # Get the current time + 1 minute
            # current_time = now + timedelta(0, 60)
            # print("Current Time =", str(current_time))
            # print("Current Time =", str(current_time)[11:13])
            # print("Current Time =", str(current_time)[14:16])
            # press ctrl + t to open a new tab

            # set in the next minute from current time
            # pywhatkit.sendwhatmsg(ph[i], message, int(str(current_time)[11:13]),int(str(current_time)[14:16]))
        # pywhatkit.sendwhatmsg_to_group(ph[0], message, 10,24)
        # for contact in ph:
        #     # get current time
        #     now = datetime.now()
        #     # get current time + 1 minute
        #     current_time = now
        #     # current_time = now
        #     print("Current Time =", str(current_time)[11:13])
        #     print("Current Time =", str(current_time)[14:16])

        #     pywhatkit.sendwhatmsg(contact, message, 10,20)#int(str(current_time)[11:13]), int(str(current_time)[14:16])+1)
        #     pg.press('enter')
        # print("Done")
        # return redirect('LabourRequest')
        from twilio.rest import Client

        account_sid = 'AC0539d2c6df9430189efa4874bef56979'
        auth_token = '1584009860dd2ad140a54dd152db9129'
        client = Client(account_sid, auth_token)

        message = client.messages.create(
            from_='+12546154637',
            body='from local host python',
            to='+919825040957'
        )

        print(message.sid)
    return render(request,'LabourReport/Management/LabourRequest.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def HomeMang(request):
    # get current time
    today = datetime.now()
    yesterday = today - timedelta(1)
    tomorrow = today + timedelta(1)
    # if time is less than 20:00:00
    if today.hour < 20 and today.hour >= 8:
        start = today.replace(hour=8, minute=0, second=0, microsecond=0)
        finish = today.replace(hour=20, minute=0, second=0, microsecond=0)
    elif today.hour >= 20:
        start = today.replace(hour=20, minute=0, second=0, microsecond=0)
        finish = tomorrow.replace(hour=8, minute=0, second=0, microsecond=0)
    elif today.hour < 8:
        start = yesterday.replace(hour=20, minute=0, second=0, microsecond=0)
        finish = today.replace(hour=8, minute=0, second=0, microsecond=0)
    data = SiteEngDay.objects.filter(created_at__range=[start,finish])
    # Convert data to dataframe
    df = pd.DataFrame(columns=['date','Areaname','ContractorName','CategoryName','LabourCategory','StructureName','NoLabor'])
    for i in data:
        dat = str(i.created_at)[:10]
        dat = dat[8:] + "-" + dat[5:7] + "-" + dat[:4]
        data_dict = {'date':dat,
                     'Areaname':str(i.Areaname),
                     'ContractorName':str(i.ContractorName),
                     'CategoryName':str(i.CategoryName),
                     'LabourCategory':str(i.LabourCategory),
                     'StructureName':str(i.StructureName),
                     'NoLabor':i.NoLabor}
        df.loc[len(df)] = data_dict

    ReportStatus = Report_Status.objects.filter(Date=today)



    Day_pie_chart = df[['date','Areaname','NoLabor']]
    Day_pie_chart = Day_pie_chart[['Areaname','NoLabor']]
    # sum the no of labors for each area
    Day_pie_chart = Day_pie_chart.groupby(['Areaname']).sum().reset_index()
    if (Day_pie_chart.empty) :
        # add column of values
        Day_pie_chart['NoLabor'] = 0

    Day_table_data = df[['Areaname','CategoryName','LabourCategory','NoLabor']]
    Day_table_data = Day_table_data.groupby(['Areaname','LabourCategory']).sum().reset_index()
    lst = ['FORM WORK','BOARD WORK','REINFORCEMENT','DISPATCH CLEARANCE','PAINTING','MASONARY']
    # MAKE OTHERS CATEGORY
    Day_table_data.loc[~Day_table_data['LabourCategory'].isin(lst), 'LabourCategory'] = 'OTHERS'
    Day_table_data = Day_table_data.groupby(['Areaname','LabourCategory']).sum().reset_index()
    if Day_table_data.empty :
        Day_table_data['NoLabor'] = 0
    Day_table_data = Day_table_data.pivot(index='Areaname',columns='LabourCategory',values='NoLabor').reset_index()

    # add all the columns from lst
    for i in lst:
        if i not in Day_table_data.columns:
            Day_table_data[i] = 0
    Day_table_data = Day_table_data.fillna(0)
    # if not(Day_table_data.empty) :

    #     Day_table_data = Day_table_data.astype({'CARPENTER':int,'BARBENDER':int,'MASON':int,'PAINTER':int,'OTHERS':int})
    # iterate over the rows

    # sequence of columns
    Day_table_data['Total'] = Day_table_data.iloc[:,1:].sum(axis=1)
    # Day_table_data = Day_table_data[['Areaname','CARPENTER','BARBENDER','MASON','PAINTER','OTHERS','Total']]


    # add a row to Day_table_data
    Day_table_data.loc[len(Day_table_data)] = Day_table_data.sum(axis=0)
    Day_table_data.iloc[-1,0] = 'Overall'

    # change name of column
    Day_table_data = Day_table_data.rename(columns={'DISPATCH CLEARANCE' : 'DISPATCHCLEARANCE','FORM WORK' : 'FORMWORK','BOARD WORK' : 'BOARDWORK'})

    ################################################################################################################################################

    if today.hour < 20 and today.hour >= 8:
        start = today.replace(hour=8, minute=0, second=0, microsecond=0)
        finish = today.replace(hour=20, minute=0, second=0, microsecond=0)
    elif today.hour >= 20:
        start = today.replace(hour=20, minute=0, second=0, microsecond=0)
        finish = tomorrow.replace(hour=8, minute=0, second=0, microsecond=0)
    elif today.hour < 8:
        start = yesterday.replace(hour=20, minute=0, second=0, microsecond=0)
        finish = today.replace(hour=8, minute=0, second=0, microsecond=0)
    print("start",start)
    print("finish",finish)
    data = SiteEngNight.objects.filter(created_at__range=[start,finish])
    for i in data:
        print(i.created_at)
    # Convert data to dataframe
    df = pd.DataFrame(columns=['date','Areaname','ContractorName','CategoryName','LabourCategory','StructureName','NoLabor'])
    for i in data:
        dat = str(i.created_at)[:10]
        dat = dat[8:] + "-" + dat[5:7] + "-" + dat[:4]
        data_dict = {'date':dat,
                     'Areaname':str(i.Areaname),
                     'ContractorName':str(i.ContractorName),
                     'CategoryName':str(i.CategoryName),
                     'LabourCategory':str(i.LabourCategory),
                     'StructureName':str(i.StructureName),
                     'NoLabor':i.NoLabor}
        df.loc[len(df)] = data_dict

    ReportStatus = Report_Status.objects.filter(Date=today)


    print(df)
    Night_pie_chart = df[['date','Areaname','NoLabor']]
    Night_pie_chart = Night_pie_chart[['Areaname','NoLabor']]
    # sum the no of labors for each area
    Night_pie_chart = Night_pie_chart.groupby(['Areaname']).sum().reset_index()
    if (Night_pie_chart.empty) :
        Night_pie_chart['NoLabor'] = 0

    Night_table_data = df[['Areaname','CategoryName','LabourCategory','NoLabor']]
    Night_table_data = Night_table_data.groupby(['Areaname','LabourCategory']).sum().reset_index()
    lst = ['FORM WORK','BOARD WORK','REINFORCEMENT','DISPATCH CLEARANCE','PAINTING','MASONARY']
    # MAKE OTHERS CATEGORY
    Night_table_data.loc[~Night_table_data['LabourCategory'].isin(lst), 'LabourCategory'] = 'OTHERS'
    Night_table_data = Night_table_data.groupby(['Areaname','LabourCategory']).sum().reset_index()
    if Night_table_data.empty :
        Night_table_data['NoLabor'] = 0
    Night_table_data = Night_table_data.pivot(index='Areaname',columns='LabourCategory',values='NoLabor').reset_index()
    # add all the columns from lst
    for i in lst:
        if i not in Night_table_data.columns:
            Night_table_data[i] = 0
    Night_table_data = Night_table_data.fillna(0)
    # if not(Night_table_data.empty) :

    #     Night_table_data = Night_table_data.astype({'CARPENTER':int,'BARBENDER':int,'MASON':int,'PAINTER':int,'OTHERS':int})
    # iterate over the rows

    # sequence of columns
    Night_table_data['Total'] = Night_table_data.iloc[:,1:].sum(axis=1)
    # Night_table_data = Night_table_data[['Areaname','CARPENTER','BARBENDER','MASON','PAINTER','OTHERS','Total']]


    # add a row to Night_table_data
    Night_table_data.loc[len(Night_table_data)] = Night_table_data.sum(axis=0)
    Night_table_data.iloc[-1,0] = 'Overall'

    # change name of column
    Night_table_data = Night_table_data.rename(columns={'DISPATCH CLEARANCE' : 'DISPATCHCLEARANCE','FORM WORK' : 'FORMWORK','BOARD WORK' : 'BOARDWORK'})


    ############################################################################################################################################################
    RS_df = pd.DataFrame(columns=['Area','Date','Status_Day','Status_Night'])
    if ReportStatus.count() == 0:
        Report_Status.objects.update(Date=today,Status_Day=False,Status_Night=False)
    for i in ReportStatus:
        dat = str(i.Date)[:10]
        dat = dat[8:] + "-" + dat[5:7] + "-" + dat[:4]
        data_dict = {'Area':str(i.Area),
                         'Date':dat,
                         'Status_Day':str(i.Status_Day),
                         'Status_Night':str(i.Status_Night)}
        RS_df.loc[len(RS_df)] = data_dict
    RS_df = RS_df.sort_values(by='Date',ascending=False)
    RS_df = RS_df.drop_duplicates(subset=['Area'],keep='first')
    RS_df = RS_df.reset_index(drop=True)

    return render(request,'LabourReport/Management/HomeMang.html',{
        # 'sli_line_chart_labels':sli_line_chart_labels,
        # 'sli_line_chart_data':sli_line_chart_data,
        # 'line_chart_labels':line_chart_labels,
        # 'line_chart_data':line_chart_data,
        'Day_table_data':Day_table_data,
        'Day_pie_chart_label':Day_pie_chart['Areaname'].values.tolist(),
        'Day_pie_chart_data':Day_pie_chart['NoLabor'].values.tolist(),
        'Night_table_data':Night_table_data,
        'Night_pie_chart_label':Night_pie_chart['Areaname'].values.tolist(),
        'Night_pie_chart_data':Night_pie_chart['NoLabor'].values.tolist(),
        'RS':RS_df,
        })

def generate_Site_report(shift,date):
        date1=datetime.strptime(date, "%Y-%m-%d")
        tomorrow = date1 + timedelta(1)
        d1=date1.strftime("%Y-%m-%d")
        d2=tomorrow.strftime("%Y-%m-%d")
        print("d",d1,d2)

        workbook = Workbook()
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Deployment Report'
        area_arr=["SBN","KV","DBM","RKP","MPZ","HBM","ALK","AIIMS","Casting Yard","Casting Yard QC","Casting Yard PM"]
        area_arr1=["Labour Category","","SBN","","KV","","DBM","","RKP","","MPZ","","HBM","","ALK","","AIIMS","","Casting Yard","","Casting Yard QC","","Casting Yard PM","","Total",""]
        area_arr2=["","","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI","SE","SLI"]
        LabCat=[]

        # Deployment Sheets
        for col_num, column_title in enumerate(area_arr1[2:], 1):

            if column_title=="":
                    # merge with next cell
                worksheet.merge_cells(start_row=1, start_column=col_num-1+2, end_row=1, end_column=col_num+2)
                #     # align to center
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell = worksheet.cell(row=1, column=col_num+2)
                cell.value = column_title
            worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
            worksheet.cell(row=1, column=1).value = "Labour Category"
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        for col_num, column_title in enumerate(area_arr2[2:], 1):
            cell = worksheet.cell(row=2, column=col_num+2)
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
        for cell in worksheet["2:2"]:
            cell.font = Font(bold=True)


        # Deployment Sheets
        if shift == 'Day':
            se_data = SiteEngDay.objects.filter(created_at__range=[d1,d2])
            sli_data = SLIDay.objects.filter(created_at__range=[d1,d2])
        elif shift == 'Night':
            se_data = SiteEngNight.objects.filter(created_at__range=[d1,d2])
            sli_data = SLINight.objects.filter(created_at__range=[d1,d2])

        df_se = pd.DataFrame(columns = ['LabourCategory', 'NoLabor','Areaname'])
        df_sli = pd.DataFrame(columns = ['LabourCategory', 'NoLabor','Areaname'])
        for i in se_data:
            dic = {'LabourCategory':str(i.LabourCategory),'Areaname':str(i.Areaname),'NoLabor':i.NoLabor}
            df_se.loc[len(df_se)] = dic
        for i in sli_data:
            dic = {'LabourCategory':str(i.LabourCategory),'Areaname':str(i.Areaname),'NoLabor':i.NoLabor}
            df_sli.loc[len(df_sli)] = dic
        df_se = df_se.groupby(['LabourCategory','Areaname']).sum().reset_index()
        df_sli = df_sli.groupby(['LabourCategory','Areaname']).sum().reset_index()

        # pivot table
        df_se = df_se.pivot_table(index=['LabourCategory'], columns=['Areaname'], values='NoLabor').reset_index()
        if df_sli.empty == False:
            df_sli = df_sli.pivot_table(index=['LabourCategory'], columns=['Areaname'], values='NoLabor').reset_index()

        df_se.columns = df_se.columns.map(lambda x: str(x) + '_se')
        df_sli.columns = df_sli.columns.map(lambda x: str(x) + '_sli')

        df_se = df_se.rename(columns={'LabourCategory_se':'LabourCategory'})
        df_sli = df_sli.rename(columns={'LabourCategory_sli':'LabourCategory'})

        df_se = df_se.fillna(0)
        df_sli = df_sli.fillna(0)
        df = pd.merge(df_se, df_sli, on='LabourCategory', how='outer')
        # arrange columns in this order ["SBN","KV","DBM","RKP","MPZ","HBM","ALK","AIIMS","Casting Yard","Casting Yard QC","Casting Yard PM"]
        df1 = pd.DataFrame(columns = ['LabourCategory', 'SBN_se', 'SBN_sli', 'KV_se', 'KV_sli', 'DBM_se', 'DBM_sli', 'RKP_se', 'RKP_sli', 'MPZ_se', 'MPZ_sli', 'HBM_se', 'HBM_sli', 'ALK_se', 'ALK_sli', 'AIIMS_se', 'AIIMS_sli', 'Casting Yard_se', 'Casting Yard_sli', 'Casting Yard QC_se', 'Casting Yard QC_sli', 'Casting Yard PM_se', 'Casting Yard PM_sli'])
        df1 = pd.concat([df1, df], axis=0)
        df1 = df1.fillna(0)

        for index, row in df1.iterrows():
            # print(index)
            worksheet.cell(row=index+3, column=1).value = row.LabourCategory
            worksheet.cell(row=index+3, column=1).alignment = Alignment(horizontal='center', vertical='center')
            # merge cells
            worksheet.merge_cells(start_row=index+3, start_column=1, end_row=index+3, end_column=2)
            # add df values to excel
            for i in range(1, len(row)):
                worksheet.cell(row=index+3, column=i+2).value = row[i]
                worksheet.cell(row=index+3, column=i+2).alignment = Alignment(horizontal='center', vertical='center')

            worksheet.cell(row=index+3, column=len(row)+2).value = '=SUM(C'+str(index+3)+'+E'+str(index+3)+'+G'+str(index+3)+'+I'+str(index+3)+'+K'+str(index+3)+'+M'+str(index+3)+'+O'+str(index+3)+'+Q'+str(index+3)+'+S'+str(index+3)+'+U'+str(index+3)+'+W'+str(index+3)+')'
            worksheet.cell(row=index+3, column=len(row)+3).value = '=SUM(D'+str(index+3)+'+F'+str(index+3)+'+H'+str(index+3)+'+J'+str(index+3)+'+L'+str(index+3)+'+N'+str(index+3)+'+P'+str(index+3)+'+R'+str(index+3)+'+T'+str(index+3)+'+V'+str(index+3)+'+X'+str(index+3)+')'
            worksheet.cell(row=index+3, column=len(row)+2).alignment = Alignment(vertical='center')
            worksheet.cell(row=index+3, column=len(row)+3).alignment = Alignment(vertical='center')
        worksheet.cell(row=index+4, column=1).value = 'Total'
        worksheet.cell(row=index+4, column=1).font = Font(bold=True)
        worksheet.cell(row=index+4, column=1).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells(start_row=index+4, start_column=1, end_row=index+4, end_column=2)
        for i in range(3,len(row)+4):
            worksheet.cell(row=index+4, column=i).value = '=SUM('+chr(64+i)+'3:'+chr(64+i)+str(index+3)+')'


        overall_dict = {}

        # Site Sheets
        for i in area_arr:
            area=i
            area_id=Area.objects.filter(AreaName=area)
            id_list = []
            for i in area_id:
                id_list.append(i.id)
            # Create multiple sheets using openpyxl
            worksheet = workbook.create_sheet(area)
            columns = ['Site Name','Contractor Name', 'Labour Category', 'Category of Deployment','Structure', 'Deployment']
            row_num = 1
                # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title

            # Define the titles for columns

            if shift == 'Day':
                rows = SiteEngDay.objects.filter(Areaname__in=id_list,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory','CategoryName','StructureName', 'NoLabor')
            elif shift == 'Night':
                rows = SiteEngNight.objects.filter(Areaname__in=id_list,created_at__range=[d1,d2]).values_list('ContractorName', 'LabourCategory','CategoryName','StructureName', 'NoLabor')

            worksheet['A1'].font = Font(bold=True)
            for cell in worksheet["1:1"]:
                cell.font = Font(bold=True)
            tot_lab=tot_help=tot_tot=0
            rows=list(rows)
            for row in rows:
                row_num += 1
                ContractorName = ContractorDetail.objects.filter(pk=row[0]).values_list('ContractorName')
                ContName=list(ContractorName[0])
                LabourCategory = LabourOfContractor.objects.filter(pk=row[1]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                LabourCategory = AddLabour.objects.filter(pk=LabourCategory[0]).values_list('LabourCategory')
                LabourCategory =list(LabourCategory[0])
                CategoryName = CategoryOfDeployment.objects.filter(pk=row[2]).values_list('CategoryName')
                CategoryName =list(CategoryName[0])
                StructureName = Structure.objects.filter(pk=row[3]).values_list('StructureName')
                StructureName =list(StructureName[0])
                row=[
                    area,
                    ContName[0],
                    LabourCategory[0],
                    CategoryName[0],
                    StructureName[0],
                    row[4],
                ]

                if row[2] not in overall_dict:
                    overall_dict[row[2]] = row[5]
                else:
                    overall_dict[row[2]] += row[5]
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value

            # write total formula in cell
            if len(rows)>0:
                worksheet['E'+str(row_num+1)] = 'Total'
                worksheet['F'+str(row_num+1)] = '=SUM(F2:F'+str(row_num)+')'
                worksheet['E'+str(row_num+1)].font = Font(bold=True)
            else:
                # merge cells
                worksheet.merge_cells('A'+str(row_num+1)+':E'+str(row_num+1))
                worksheet['A'+str(row_num+1)] = 'No Data Found'
                # align to center
                worksheet['A'+str(row_num+1)].alignment = Alignment(horizontal='center')

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length+3
                worksheet.column_dimensions[column].width = adjusted_width

        worksheet = workbook.create_sheet('Overall')
        overall_se = SiteEngDay.objects.filter(created_at__range=[d1,d2])
        overall_sli = SLIDay.objects.filter(created_at__range=[d1,d2])

        row_num = 1
        columns = ['Labour Category', 'SE-Deployment', 'SLI-Deployment']

        worksheet['A1'] = 'Labour Category'
        worksheet['B1'] = 'SE-Deployment'
        worksheet['C1'] = 'SLI-Deployment'
        worksheet['A1'].font = Font(bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['B1'].font = Font(bold=True)
        worksheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['C1'].font = Font(bold=True)
        worksheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

        # Convert overall_se to dataframe
        df_se = pd.DataFrame(columns = ['LabourCategory', 'SE_NoLabor'])
        for i in overall_se:
            dic = {'LabourCategory':str(i.LabourCategory),'SE_NoLabor':i.NoLabor}
            df_se.loc[len(df_se)] = dic
        df_sli = pd.DataFrame(columns = ['LabourCategory', 'SLI_NoLabor'])
        for i in overall_sli:
            dic = {'LabourCategory':str(i.LabourCategory),'SLI_NoLabor':i.NoLabor}
            df_sli.loc[len(df_sli)] = dic
        df_se = df_se.groupby(['LabourCategory']).sum().reset_index()
        df_sli = df_sli.groupby(['LabourCategory']).sum().reset_index()
        df = pd.merge(df_se, df_sli, on='LabourCategory', how='outer')
        df = df.fillna(0)
        for index, row in df.iterrows():
            row_num += 1
            if len(row) == 3:
                row = [
                    row[0],
                    row[1],
                    row[2],
                ]
            else:
                row = [
                    row[0],
                    row[1],
                    0,
                ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        # write total formula in cell
        worksheet['A'+str(row_num+1)] = 'Total'
        worksheet['B'+str(row_num+1)] = '=SUM(B2:B'+str(row_num)+')'
        worksheet['C'+str(row_num+1)] = '=SUM(C2:C'+str(row_num)+')'
        worksheet['A'+str(row_num+1)].font = Font(bold=True)
        worksheet['A'+str(row_num+1)].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['B'+str(row_num+1)].font = Font(bold=True)
        worksheet['B'+str(row_num+1)].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['C'+str(row_num+1)].font = Font(bold=True)
        worksheet['C'+str(row_num+1)].alignment = Alignment(horizontal='center', vertical='center')

        for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length+3
                worksheet.column_dimensions[column].width = adjusted_width
        return workbook

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Site Engineer'])
def Change_Report_Status(request,shift):
    print("Inside Change_Report_Status")
    current_user = request.user
    Areaname = Area.objects.filter(Username=current_user.username)
    Areaname_id=Areaname[0].id
    Areaname = Areaname[0].AreaName
    # get the report status
    report_status_all = Report_Status.objects.all()
    # report_status = report_status[0]
    print(report_status_all,Areaname,shift)

    # add data to report_status
    if shift == 'Day':
        dic = {
            'Area' : Areaname,
            'Date' : datetime.now(),
            'Status_Day' : True,
        }
        report_status = Report_Status.objects.filter(Area=Areaname)
        if report_status:
            report_status = report_status[0]
            report_status.Date = datetime.now()
            if report_status.Status_Day == True:
                report_status.Status_Day = False
            else:
                report_status.Status_Day = True
            report_status.save()
        else:
            report_status = Report_Status.objects.create(**dic)
    elif shift == 'Night':
        dic = {
            'Area' : Areaname,
            'Date' : datetime.now(),
            'Status_Night' : True,
        }
        report_status = Report_Status.objects.filter(Area=Areaname)
        if report_status:
            report_status = report_status[0]
            report_status.Date = datetime.now()
            if report_status.Status_Night == True:
                report_status.Status_Night = False
            else:
                report_status.Status_Night = True
            report_status.save()
        else:
            report_status = Report_Status.objects.create(**dic)
    # get all the Area name from Report_Status
    report_status_area = Report_Status.objects.all().values_list('Area',flat=True)

    if CheckStatusAllArea('Day'):
        Send_Emails('Day')


    return redirect('Login')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def SiteReport(request):
    current_user = request.user
    if request.method == 'POST':
        shift = request.POST.get('shift')
        date = request.POST.get('date')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename="Deployment Report "+date+" "+shift+".xlsx"
        response['Content-Disposition'] = 'attachment; filename='+filename
        workbook = generate_Site_report(shift,date)
        workbook.save(response)
        return response

    return render(request,'LabourReport/Management/SiteReport.html')

@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def FinalReport(request):
    if request.method == 'POST':
        shift = request.POST.get('shift')
        date = request.POST.get('date')
        date1=datetime.strptime(date, "%Y-%m-%d")
        tomorrow = date1 + timedelta(1)
        d1=date1.strftime("%Y-%m-%d")
        d2=tomorrow.strftime("%Y-%m-%d")
        reporttype = request.POST.get('type')

        print(shift,date,d1,d2,reporttype)

        workbook = Workbook()


        if reporttype == 'Site':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename="Site Report "+date+" "+shift+".xlsx"
            response['Content-Disposition'] = 'attachment; filename='+filename
            worksheet = workbook.active
            worksheet.title = 'Site Report'+date+' '+shift
            area=["SBN","KV","DBM","RKP","MPZ","HBM","ALK","AIIMS","Casting Yard","Casting Yard QC","Casting Yard PM"]

            columns = ['Contractor Name', 'Labour Category', 'Category of Deployment','Structure', 'Deployment']
            worksheet.column_dimensions['A'].width = len(columns[0])
            worksheet.column_dimensions['B'].width = len(columns[1])

            worksheet['A1'] = 'Contractor Name'
            worksheet['B1'] = 'Types of Labour'
            cureent_cell=worksheet.cell(row=1,column=ord('A')-64)
            cureent_cell.alignment = Alignment(horizontal='center', vertical='center')
            cureent_cell=worksheet.cell(row=1,column=ord('B')-64)
            cureent_cell.alignment = Alignment(horizontal='center', vertical='center')

        workbook.save(response)
        # return response
    return render(request,'LabourReport/Management/FinalReport.html')


def generate_email_report():
    today = datetime.now()
    start_day = today.replace(hour=8, minute=0, second=0, microsecond=0)
    end_day = today.replace(hour=20, minute=0, second=0, microsecond=0)
    yesterday = today - timedelta(1)
    start_night = yesterday.replace(hour=20, minute=0, second=0, microsecond=0)
    end_night = today.replace(hour=8, minute=0, second=0, microsecond=0)
    data_Day = SiteEngDay.objects.filter(created_at__range=[start_day,end_day])
    data_Night = SiteEngNight.objects.filter(created_at__range=[start_night,end_night])

    # Convert data to dataframe
    df = pd.DataFrame(columns=['Shift','date','Areaname','ContractorName','CategoryName','LabourCategory','StructureName','NoLabor'])
    for i in data_Day:
        dat = str(i.created_at)[:10]
        dat = dat[8:] + "-" + dat[5:7] + "-" + dat[:4]
        data_dict = {'Shift':'Day',
                     'date':dat,
                     'Areaname':str(i.Areaname),
                     'ContractorName':str(i.ContractorName),
                     'CategoryName':str(i.CategoryName),
                     'LabourCategory':str(i.LabourCategory),
                     'StructureName':str(i.StructureName),
                     'NoLabor':i.NoLabor}
        df.loc[len(df)] = data_dict

    for i in data_Night:
        dat = str(i.created_at)[:10]
        dat = dat[8:] + "-" + dat[5:7] + "-" + dat[:4]
        data_dict = {'Shift':'Night',
                     'date':dat,
                     'Areaname':str(i.Areaname),
                     'ContractorName':str(i.ContractorName),
                     'CategoryName':str(i.CategoryName),
                     'LabourCategory':str(i.LabourCategory),
                     'StructureName':str(i.StructureName),
                     'NoLabor':i.NoLabor}
        df.loc[len(df)] = data_dict

    table_data =df
    # in Areaname replace 'Casting Yard QC','Casting Yard PM' and 'Casting Yard' with 'Casting Yard Work'
    table_data.loc[table_data['Areaname'].isin(['Casting Yard QC','Casting Yard PM','Casting Yard']),'Areaname'] = 'Casting Yard Work'
    # make other Areaname as 'Station Work'
    table_data.loc[~table_data['Areaname'].isin(['Casting Yard Work']),'Areaname'] = 'Station Work'

    # table_data['CategoryName'] = table_data['Shift'] + ' ' + table_data['CategoryName']

    # if
    table_data = table_data.groupby(['ContractorName','LabourCategory','CategoryName','Shift']).sum().reset_index()

    table_data = table_data.pivot_table(index=['ContractorName','LabourCategory'],
                                        columns=['CategoryName','Shift'],values='NoLabor')


    # for i in lst:
    #     if i not in table_data.columns:
    #             print(i,"not present")
    #             table_data[i] = 0
    # table_data = table_data[lst]

    table_data = table_data.fillna(0)

    # add columns with shift equals to Day and Night there is no Day or Night in column name
    for i in table_data.columns:
        if ('Day' not in i) and ('Night' not in i):
            table_data[i,'Day'] = 0
            table_data[i,'Night'] = 0
    table_data['Day Total'] = 0
    table_data['Night Total'] = 0
    for i in table_data.columns:
        if 'Day' in i:
            table_data['Day Total'] = table_data['Day Total'] + table_data[i]
        if 'Night' in i:
            table_data['Night Total'] = table_data['Night Total'] + table_data[i]



    # table_data['Day Total'] = table_data['Day MASON'] + table_data['Day HELPER (MASON)'] + table_data['Day CARPENTER'] + table_data['Day HELPER (CARP)'] + table_data['Day BARBENDER'] + table_data['Day HELPER (BAR)'] + table_data['Day GAS CUTTER'] + table_data['Day WELDER'] + table_data['Day HELPER (WELDER)'] + table_data['Day FITTER'] + table_data['Day HELPER (FITTER)'] + table_data['Day STAGGING'] + table_data['Day RIGGER'] + table_data['Day PAINTER']  + table_data['Day ELECTRICIAN'] + table_data['Day PLUMBER'] + table_data['Day MECHANIC'] + table_data['Day SUB CONTRACTOR'] + table_data['Day OTHERS']
    # table_data['Night Total'] = table_data['Night MASON'] + table_data['Night HELPER (MASON)'] + table_data['Night CARPENTER'] + table_data['Night HELPER (CARP)'] + table_data['Night BARBENDER'] + table_data['Night HELPER (BAR)'] + table_data['Night GAS CUTTER'] + table_data['Night WELDER'] + table_data['Night HELPER (WELDER)'] + table_data['Night FITTER'] + table_data['Night HELPER (FITTER)'] + table_data['Night STAGGING'] + table_data['Night RIGGER'] + table_data['Night PAINTER'] + table_data['Night ELECTRICIAN'] + table_data['Night PLUMBER'] + table_data['Night MECHANIC'] + table_data['Night SUB CONTRACTOR'] + table_data['Night OTHERS']

    table_data['ContractorName'] = table_data.index
    table_data['LabourCategory'] = table_data['ContractorName'].astype(str).str.split(',').str[1]
    table_data['ContractorName'] = table_data['ContractorName'].astype(str).str.split(',').str[0]
    # remove ' from ContractorName
    table_data['ContractorName'] = table_data['ContractorName'].str.replace("'",'')
    table_data['ContractorName'] = table_data['ContractorName'].str.replace("(",'')
    table_data['ContractorName'] = table_data['ContractorName'].str.replace(")",'')
    table_data['LabourCategory'] = table_data['LabourCategory'].str.replace("'",'')
    table_data['LabourCategory'] = table_data['LabourCategory'].str.replace("(",'')
    table_data['LabourCategory'] = table_data['LabourCategory'].str.replace(")",'')

    table_data = table_data.reset_index(drop=True)

    # Change Index column name to Sr. No.
    table_data['Sr. No.'] = table_data.index + 1
    # Keep Sr. No., ContractorName, and Labour Category column at first position
    cols = table_data.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    table_data = table_data[cols]
    cols = table_data.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    table_data = table_data[cols]
    cols = table_data.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    table_data = table_data[cols]

    # add total row with sum of all columns except ContractorName, LabourCategory and Sr. No.
    table_data.loc['Total'] = table_data.sum(numeric_only=True, axis=0)
    # table_data = table_data.append(table_data.sum(numeric_only=True), ignore_index=True)

    # rotate first three columns with each other
    cols = table_data.columns.tolist()
    cols = cols[2:3] + cols[:2] + cols[3:]
    table_data = table_data[cols]
    # Now add Total in ContractorName column
    # table_data.loc['Total','ContractorName'] = 'Total'
    table_data = table_data.reset_index(drop=True)
    # print all the when we print table_data

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    col = len(table_data.columns)

    ws['A1'] = "Date : " + str(today)[8:10] + "." + str(today)[5:7] + "." + str(today)[:4]
    ws.merge_cells('A1:C1')
    # row = 2




    last = ''
    for i in range(0,len(table_data.columns)):
        count = 0
        for j in table_data.columns[i]:
            count += 1
            if count == 1:
                ws.cell(row=2, column=i+1).value = j
                ws.cell(row=2, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=2, column=i+1).font = Font(bold=True)
                if j == last :
                    ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=i+1)
                elif i>3 and j!=last:
                    # set width of column
                    max_length = 0

            elif count == 2:
                ws.cell(row=3, column=i+1).value = j
                ws.cell(row=3, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=3, column=i+1).font = Font(bold=True)
                if i != 0:
                    last = table_data.columns[i][0]

    ws.merge_cells('A2:A3')
    ws.merge_cells('B2:B3')
    ws.merge_cells('C2:C3')
    ws.merge_cells(start_row=2, start_column=len(table_data.columns), end_row=3, end_column=len(table_data.columns))
    ws.merge_cells(start_row=2, start_column=len(table_data.columns)-1, end_row=3, end_column=len(table_data.columns)-1)
            # ws.cell(row=4, column=i).value = i

    for r in dataframe_to_rows(table_data, index=False, header=False):
        ws.append(r)

    ws.delete_rows(5)

    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in range(1,ws.max_column+1):
        row = ws.max_row
        ws.cell(row=row,column=col).font = Font(bold=True)

    for cell in ws["A:A"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws["B:B"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws["C:C"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in range(ws.max_column-1,ws.max_column+1):
        for row in range(2,len(table_data)+3):
            ws.cell(row=row,column=col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row,column=col).font = Font(bold=True)

    max_length = 0
    for cell in ws["B"]:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length
    ws.column_dimensions['B'].width = adjusted_width

    max_length = 0
    for cell in ws["C"]:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length
    ws.column_dimensions['C'].width = adjusted_width

    # change background color of cells
    for col in range(1,ws.max_column+1):
        ws.cell(row=1,column=col).fill = PatternFill(start_color='FCE4D6',end_color='FCE4D6',fill_type='solid')

    for col in range(1,ws.max_column+1):
        for row in range(2,4):
            ws.cell(row=row,column=col).fill = PatternFill(start_color='B4C6E7',end_color='B4C6E7',fill_type='solid')

    ws.cell(row=len(table_data)+2,column=1).value = 'TOTAL MANPOWER'
    ws.merge_cells(start_row=len(table_data)+2, start_column=1, end_row=len(table_data)+2, end_column=3)
    ws.cell(row=len(table_data)+2,column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=len(table_data)+2,column=1).font = Font(bold=True)

    for cell in ws[str(len(table_data)+2)+":"+str(len(table_data)+2)]:
        cell.font = Font(bold=True)

    # border for cells
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    for col in range(1,ws.max_column+1):
        for row in range(2,ws.max_row+1):
            ws.cell(row=row,column=col).border = thin_border

    # freeze first four row
    ws.freeze_panes = ws['D4']

    return wb


@login_required(login_url='Login')
@allowed_users(allowed_roles=['Management'])
def Send_Emails(request,shift):
        # get today date
        today = date.today()
        today = today.strftime("%Y-%m-%d")
        today = str(today)[:10]
        print(str(today)[:10])
        workbook = generate_email_report()
        print("Workbook")
        excelfile = BytesIO()
        workbook.save(excelfile)
        email = EmailMessage(
            'Deployment Report '+ today,
            'Please find the attached Deployment Report,',
            settings.EMAIL_HOST_USER,
            ['harshit.kava@gmail.com','atulkava@gmail.com','gokulakrishnan.s@urcc.co.in','prakash.g@urcc.co.in','priya.r@urcc.co.in','dineshbabu.s@urcc.co.in'],
            )

        email.attach('Deployment Report '+ today +'.xlsx',excelfile.getvalue(),'application/ms-excel')
        email.fail_silently = False
        email.send()
        print("Email Sent")
        return redirect('HomeMang')

# @login_required(login_url='Login')
# @allowed_users(allowed_roles=['Management'])
# def RFI(request):
#     form = RFI_FORM()
#     if request.method == 'POST':
#         print(request.FILES)
#         form = RFI_FORM(request.POST,request.FILES)
#         if form.is_valid():
#             print("valid")
#             form.save()
#             return redirect('RFI')
#     return render(request,'LabourReport/Management/RFI.html',{'form':form})